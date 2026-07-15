# ============================================================
# MODELLO PROFILAZIONE CLIENTI IN SCONFINO
# Cluster30 / Cluster90 / Cluster90+
#
# Versione V3:
# - rimozione leakage
# - profiling cluster
# - test statistici
# - decision tree
# - random forest
# - gradient boosting
# - feature importance
# - score 0-100
# - cliente tipo
# - business insight automatici
# - export Excel finale
# - salvataggio pipeline Random Forest
# - salvataggio metadati e profili di training
# - modalità TRAIN / PREDICT
# - scoring di nuovi dataset senza target
# - confronto distribuzione cluster training vs nuovo dataset
# - controllo drift sulle variabili numeriche
# - valutazione del nuovo dataset se il target è disponibile
# ============================================================

import json
import warnings
from pathlib import Path

import joblib
import numpy as np
import pandas as pd

from scipy.stats import chi2_contingency, kruskal

from sklearn.compose import ColumnTransformer
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.impute import SimpleImputer
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.model_selection import cross_val_score, train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.tree import DecisionTreeClassifier, export_text

warnings.filterwarnings("ignore")


# ============================================================
# 1. PARAMETRI
# ============================================================

# Valori ammessi:
# - "TRAIN": addestra, valuta, salva il modello e produce il report storico
# - "PREDICT": carica il modello già salvato e valuta un nuovo dataset
MODE = "TRAIN"

# File usato in modalità TRAIN
INPUT_FILE = "dataset_clienti_cluster.xlsx"

# File usato in modalità PREDICT
NEW_DATA_FILE = "nuovo_dataset_clienti.xlsx"

# Output
TRAIN_OUTPUT_FILE = "output_modello_cluster_gestione_v3.xlsx"
PREDICT_OUTPUT_FILE = "output_nuovo_dataset_cluster_v3.xlsx"

# Cartella degli artefatti persistenti
MODEL_DIR = Path("model_cluster_gestione")
MODEL_FILE = MODEL_DIR / "random_forest_pipeline.joblib"
METADATA_FILE = MODEL_DIR / "model_metadata.json"
TRAINING_PROFILE_FILE = MODEL_DIR / "training_profiles.xlsx"

TARGET_COL = "cluster"

VALID_CLUSTERS = ["cluster30", "cluster90", "cluster90+"]

CLUSTER_RANK = {
    "cluster30": 30,
    "cluster90": 60,
    "cluster90+": 100
}

LEAKAGE_COLS = [
    "gestione",
    "giorni_gestione",
    "max_giorni_gestione",
    "num_gestioni",
    "giorni_totali_gestione"
]

ID_COLS = [
    "NDG",
    "NDG2",
    "ndg",
    "ndg2",
    "ID_CLIENTE",
    "id_cliente"
]

TARGET_DERIVED_COLS = [
    "target_ordinale"
]

# Soglie drift.
# Una variabile è segnalata se:
# 1. la variazione relativa della mediana supera questa soglia;
# 2. oppure la differenza standardizzata supera DRIFT_STD_THRESHOLD.
DRIFT_RELATIVE_THRESHOLD = 0.20
DRIFT_STD_THRESHOLD = 0.50

RANDOM_STATE = 42


# ============================================================
# 2. FUNZIONI GENERALI
# ============================================================

def normalize_target(series):
    """Normalizza le etichette del target nei tre cluster previsti."""
    normalized = series.astype(str).str.lower().str.strip()

    return normalized.replace({
        "cluster 30": "cluster30",
        "cluster_30": "cluster30",
        "30": "cluster30",

        "cluster 90": "cluster90",
        "cluster_90": "cluster90",
        "90": "cluster90",

        "cluster 90+": "cluster90+",
        "cluster_90+": "cluster90+",
        "cluster90plus": "cluster90+",
        "cluster 90 plus": "cluster90+",
        "cluster180": "cluster90+",
        "cluster180+": "cluster90+",
        "180": "cluster90+",
        "90+": "cluster90+"
    })


def safe_relative_change(new_value, old_value):
    """
    Calcola la variazione relativa evitando divisioni instabili per valori
    storici nulli o molto vicini a zero.
    """
    if pd.isna(new_value) or pd.isna(old_value):
        return np.nan

    if abs(old_value) < 1e-9:
        if abs(new_value) < 1e-9:
            return 0.0
        return np.nan

    return (new_value - old_value) / abs(old_value)


def make_preprocessor(numeric_features, categorical_features):
    """Crea il preprocessing completo da salvare insieme al modello."""
    numeric_transformer = Pipeline(steps=[
        ("imputer", SimpleImputer(strategy="median")),
        ("scaler", StandardScaler())
    ])

    categorical_transformer = Pipeline(steps=[
        ("imputer", SimpleImputer(strategy="most_frequent")),
        ("encoder", OneHotEncoder(handle_unknown="ignore"))
    ])

    transformers = []

    if numeric_features:
        transformers.append(("num", numeric_transformer, numeric_features))

    if categorical_features:
        transformers.append(("cat", categorical_transformer, categorical_features))

    if not transformers:
        raise ValueError("Nessuna variabile utilizzabile trovata nel dataset.")

    return ColumnTransformer(
        transformers=transformers,
        remainder="drop"
    )


def get_feature_names(fitted_preprocessor, numeric_features, categorical_features):
    """Recupera i nomi delle feature anche con versioni sklearn meno recenti."""
    feature_names = []

    if numeric_features:
        feature_names.extend(numeric_features)

    if categorical_features:
        cat_encoder = (
            fitted_preprocessor
            .named_transformers_["cat"]
            .named_steps["encoder"]
        )

        try:
            cat_names = cat_encoder.get_feature_names_out(categorical_features)
        except AttributeError:
            cat_names = cat_encoder.get_feature_names(categorical_features)

        feature_names.extend(list(cat_names))

    return feature_names


def extract_id_columns(df):
    """Restituisce le colonne identificative presenti nel dataset."""
    return [col for col in ID_COLS if col in df.columns]


def build_score_dataframe(model, X, source_df, actual_target=None):
    """
    Calcola:
    - probabilità per cluster;
    - cluster predetto;
    - score ordinale 0-100;
    - fascia di persistenza;
    - eventuale cluster reale.
    """
    classifier = model.named_steps["model"]
    classes = list(classifier.classes_)

    probabilities = model.predict_proba(X)

    result = pd.DataFrame(
        probabilities,
        columns=[f"prob_{c}" for c in classes],
        index=X.index
    )

    score = np.zeros(len(result))

    for cluster in classes:
        score += (
            result[f"prob_{cluster}"].to_numpy()
            * CLUSTER_RANK.get(cluster, 0)
        )

    result["management_persistence_score"] = score
    result["predicted_cluster"] = model.predict(X)

    if actual_target is not None:
        result["actual_cluster"] = actual_target.values

    for id_col in extract_id_columns(source_df):
        result[id_col] = source_df.loc[X.index, id_col].values

    result["score_band"] = pd.cut(
        result["management_persistence_score"],
        bins=[0, 40, 70, 100],
        labels=[
            "Low persistence",
            "Medium persistence",
            "High persistence"
        ],
        include_lowest=True
    )

    # Porta gli identificativi all'inizio.
    leading_cols = extract_id_columns(source_df)

    ordered_cols = (
        leading_cols
        + ["predicted_cluster", "management_persistence_score", "score_band"]
        + [c for c in result.columns if c.startswith("prob_")]
    )

    if actual_target is not None:
        ordered_cols.append("actual_cluster")

    remaining_cols = [c for c in result.columns if c not in ordered_cols]

    return result[ordered_cols + remaining_cols].reset_index(drop=True)


def format_distribution(series, name):
    """Crea conteggi e percentuali per una distribuzione di cluster."""
    counts = series.value_counts().reindex(VALID_CLUSTERS, fill_value=0)
    percentages = counts / max(counts.sum(), 1)

    return pd.DataFrame({
        "cluster": VALID_CLUSTERS,
        f"{name}_count": counts.values,
        f"{name}_percentage": percentages.values
    })


def generate_portfolio_summary(distribution_comparison, drift_df):
    """Genera frasi manageriali sintetiche sul nuovo portafoglio."""
    summary = []

    if not distribution_comparison.empty:
        dominant_row = distribution_comparison.sort_values(
            "new_percentage",
            ascending=False
        ).iloc[0]

        summary.append(
            f"Il cluster prevalente nel nuovo dataset è "
            f"{dominant_row['cluster']} "
            f"({dominant_row['new_percentage']:.1%} dei clienti)."
        )

        largest_change = distribution_comparison.iloc[
            distribution_comparison["percentage_point_change"].abs().argmax()
        ]

        direction = (
            "aumenta"
            if largest_change["percentage_point_change"] > 0
            else "diminuisce"
        )

        summary.append(
            f"La variazione più rilevante rispetto al training riguarda "
            f"{largest_change['cluster']}: la quota {direction} di "
            f"{abs(largest_change['percentage_point_change']):.1%} punti percentuali."
        )

    if drift_df is not None and not drift_df.empty:
        flagged = drift_df[drift_df["drift_flag"] == True]

        if flagged.empty:
            summary.append(
                "Non emergono segnali rilevanti di drift sulle variabili numeriche "
                "secondo le soglie configurate."
            )
        else:
            top_drift = flagged.sort_values(
                ["abs_standardized_median_shift", "abs_relative_median_change"],
                ascending=False
            ).head(5)

            vars_text = ", ".join(top_drift["variable"].astype(str).tolist())

            summary.append(
                f"Sono state segnalate {len(flagged)} variabili con possibile drift. "
                f"Le principali sono: {vars_text}."
            )

    return pd.DataFrame({
        "portfolio_summary": summary
    })


def style_excel_file(file_path):
    """Applica una formattazione essenziale al file Excel prodotto."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(file_path)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            max_length = 0
            col_idx = column_cells[0].column

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(file_path)


# ============================================================
# 3. TRAINING
# ============================================================

def train_model():
    MODEL_DIR.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(INPUT_FILE)

    print("Righe dataset:", len(df))
    print("Colonne iniziali:", len(df.columns))

    if TARGET_COL not in df.columns:
        raise ValueError(
            f"La colonna target '{TARGET_COL}' non è presente in {INPUT_FILE}."
        )

    # --------------------------------------------------------
    # 3.1 Normalizzazione target
    # --------------------------------------------------------

    df[TARGET_COL] = normalize_target(df[TARGET_COL])
    df = df[df[TARGET_COL].isin(VALID_CLUSTERS)].copy()

    if df.empty:
        raise ValueError(
            "Dopo la normalizzazione non risultano osservazioni con cluster validi."
        )

    df["target_ordinale"] = df[TARGET_COL].map(CLUSTER_RANK)

    print("\nDistribuzione cluster:")
    print(df[TARGET_COL].value_counts())

    if df[TARGET_COL].nunique() < 2:
        raise ValueError("Il training richiede almeno due cluster distinti.")

    # --------------------------------------------------------
    # 3.2 Rimozione leakage e identificativi
    # --------------------------------------------------------

    drop_cols = (
        [TARGET_COL]
        + TARGET_DERIVED_COLS
        + LEAKAGE_COLS
        + ID_COLS
    )

    X = df.drop(columns=drop_cols, errors="ignore")
    y = df[TARGET_COL].copy()

    X = X.dropna(axis=1, how="all")

    # Elimina colonne costanti: non forniscono informazione e possono creare
    # artefatti nel preprocessing.
    constant_cols = [
        col for col in X.columns
        if X[col].nunique(dropna=False) <= 1
    ]

    X = X.drop(columns=constant_cols, errors="ignore")

    print("\nColonne dopo pulizia:", len(X.columns))

    if X.shape[1] == 0:
        raise ValueError("Nessuna feature disponibile dopo la pulizia.")

    # --------------------------------------------------------
    # 3.3 Numeric / categoric
    # --------------------------------------------------------

    numeric_features = X.select_dtypes(
        include=["number"]
    ).columns.tolist()

    categorical_features = [
        col for col in X.columns
        if col not in numeric_features
    ]

    print("\nVariabili numeriche:", len(numeric_features))
    print("Variabili categoriche:", len(categorical_features))

    # --------------------------------------------------------
    # 3.4 Profiling cluster
    # --------------------------------------------------------

    if numeric_features:
        profile_numeric = df.groupby(TARGET_COL)[numeric_features].agg([
            "count", "mean", "median", "std", "min", "max"
        ])

        cluster_median = (
            df.groupby(TARGET_COL)[numeric_features]
            .median()
            .T
        )

        cluster_mean = (
            df.groupby(TARGET_COL)[numeric_features]
            .mean()
            .T
        )
    else:
        profile_numeric = pd.DataFrame()
        cluster_median = pd.DataFrame()
        cluster_mean = pd.DataFrame()

    for cluster in VALID_CLUSTERS:
        if cluster not in cluster_median.columns:
            cluster_median[cluster] = np.nan

    if not cluster_median.empty:
        cluster_profile = cluster_median[VALID_CLUSTERS].copy()
        cluster_profile["delta_90plus_vs_30"] = (
            cluster_profile["cluster90+"]
            - cluster_profile["cluster30"]
        )
        cluster_profile["abs_delta_90plus_vs_30"] = (
            cluster_profile["delta_90plus_vs_30"].abs()
        )
    else:
        cluster_profile = pd.DataFrame()

    # Profilo globale usato successivamente per il drift.
    training_numeric_profile = pd.DataFrame()

    if numeric_features:
        training_numeric_profile = pd.DataFrame({
            "variable": numeric_features,
            "training_count": [df[col].notna().sum() for col in numeric_features],
            "training_mean": [df[col].mean() for col in numeric_features],
            "training_median": [df[col].median() for col in numeric_features],
            "training_std": [df[col].std() for col in numeric_features],
            "training_p25": [df[col].quantile(0.25) for col in numeric_features],
            "training_p75": [df[col].quantile(0.75) for col in numeric_features],
            "training_missing_rate": [df[col].isna().mean() for col in numeric_features]
        })

    # --------------------------------------------------------
    # 3.5 Test statistici
    # --------------------------------------------------------

    stat_tests = []

    for col in numeric_features:
        groups = [
            df.loc[df[TARGET_COL] == cluster, col].dropna()
            for cluster in VALID_CLUSTERS
            if cluster in df[TARGET_COL].unique()
        ]

        groups = [group for group in groups if len(group) > 5]

        if len(groups) >= 2:
            try:
                _, p_value = kruskal(*groups)
                stat_tests.append({
                    "variable": col,
                    "type": "numeric",
                    "test": "kruskal",
                    "p_value": p_value
                })
            except Exception:
                pass

    for col in categorical_features:
        try:
            table = pd.crosstab(df[TARGET_COL], df[col])

            if table.shape[0] > 1 and table.shape[1] > 1:
                _, p_value, _, _ = chi2_contingency(table)

                stat_tests.append({
                    "variable": col,
                    "type": "categorical",
                    "test": "chi_square",
                    "p_value": p_value
                })
        except Exception:
            pass

    stat_tests_df = pd.DataFrame(stat_tests)

    if not stat_tests_df.empty:
        stat_tests_df = stat_tests_df.sort_values("p_value")
        stat_tests_df["significant_005"] = (
            stat_tests_df["p_value"] < 0.05
        )
        stat_tests_df["significance"] = pd.cut(
            stat_tests_df["p_value"],
            bins=[-1, 0.001, 0.01, 0.05, 1],
            labels=["***", "**", "*", ""]
        )

    # --------------------------------------------------------
    # 3.6 Preprocessing e split
    # --------------------------------------------------------

    preprocessor = make_preprocessor(
        numeric_features=numeric_features,
        categorical_features=categorical_features
    )

    class_counts = y.value_counts()
    min_class_count = int(class_counts.min())

    if min_class_count < 2:
        raise ValueError(
            "Ogni cluster deve contenere almeno due osservazioni "
            "per effettuare train/test split stratificato."
        )

    X_train, X_test, y_train, y_test = train_test_split(
        X,
        y,
        test_size=0.25,
        random_state=RANDOM_STATE,
        stratify=y
    )

    # --------------------------------------------------------
    # 3.7 Modelli
    # --------------------------------------------------------

    tree_model = Pipeline(steps=[
        ("preprocessor", preprocessor),
        ("model", DecisionTreeClassifier(
            max_depth=4,
            min_samples_leaf=80,
            class_weight="balanced",
            random_state=RANDOM_STATE
        ))
    ])

    rf_model = Pipeline(steps=[
        ("preprocessor", preprocessor),
        ("model", RandomForestClassifier(
            n_estimators=400,
            max_depth=8,
            min_samples_leaf=40,
            class_weight="balanced",
            random_state=RANDOM_STATE,
            n_jobs=-1
        ))
    ])

    gb_model = Pipeline(steps=[
        ("preprocessor", preprocessor),
        ("model", GradientBoostingClassifier(
            n_estimators=250,
            learning_rate=0.04,
            max_depth=3,
            random_state=RANDOM_STATE
        ))
    ])

    models = {
        "Decision Tree": tree_model,
        "Random Forest": rf_model,
        "Gradient Boosting": gb_model
    }

    model_reports = {}
    confusion_matrices = {}

    for name, model in models.items():
        model.fit(X_train, y_train)
        predictions = model.predict(X_test)

        print(f"\n=== {name.upper()} ===")
        print(classification_report(
            y_test,
            predictions,
            zero_division=0
        ))

        classes = list(model.named_steps["model"].classes_)

        cm = confusion_matrix(
            y_test,
            predictions,
            labels=classes
        )

        print(cm)

        model_reports[name] = classification_report(
            y_test,
            predictions,
            output_dict=True,
            zero_division=0
        )

        confusion_matrices[name] = pd.DataFrame(
            cm,
            index=[f"actual_{c}" for c in classes],
            columns=[f"pred_{c}" for c in classes]
        )

    # --------------------------------------------------------
    # 3.8 Cross validation Random Forest
    # --------------------------------------------------------

    cv_folds = min(5, min_class_count)

    cv_scores = cross_val_score(
        rf_model,
        X,
        y,
        cv=cv_folds,
        scoring="f1_weighted"
    )

    cv_summary = pd.DataFrame({
        "metric": [
            "cv_folds",
            "cv_f1_weighted_mean",
            "cv_f1_weighted_std"
        ],
        "value": [
            cv_folds,
            cv_scores.mean(),
            cv_scores.std()
        ]
    })

    # --------------------------------------------------------
    # 3.9 Refit finale Random Forest sull'intero dataset
    # --------------------------------------------------------

    # La valutazione sopra rimane out-of-sample.
    # Per il modello operativo si esegue poi il fit su tutto il dataset,
    # così da utilizzare tutte le osservazioni disponibili.
    rf_model.fit(X, y)

    # --------------------------------------------------------
    # 3.10 Feature importance
    # --------------------------------------------------------

    rf_preprocessor = rf_model.named_steps["preprocessor"]
    rf_classifier = rf_model.named_steps["model"]

    feature_names = get_feature_names(
        rf_preprocessor,
        numeric_features,
        categorical_features
    )

    feature_importance = pd.DataFrame({
        "feature": feature_names,
        "importance": rf_classifier.feature_importances_
    }).sort_values("importance", ascending=False)

    top_features = feature_importance.head(30)

    # --------------------------------------------------------
    # 3.11 Score training
    # --------------------------------------------------------

    score_df = build_score_dataframe(
        model=rf_model,
        X=X,
        source_df=df,
        actual_target=y
    )

    # --------------------------------------------------------
    # 3.12 Decision tree rules
    # --------------------------------------------------------

    # Per produrre regole coerenti con tutto il dataset si rifitta anche
    # il decision tree sul campione completo.
    tree_model.fit(X, y)

    tree_classifier = tree_model.named_steps["model"]
    tree_preprocessor = tree_model.named_steps["preprocessor"]

    tree_feature_names = get_feature_names(
        tree_preprocessor,
        numeric_features,
        categorical_features
    )

    tree_rules = export_text(
        tree_classifier,
        feature_names=list(tree_feature_names)
    )

    print("\n=== REGOLE DECISION TREE ===")
    print(tree_rules)

    # --------------------------------------------------------
    # 3.13 Cliente tipo
    # --------------------------------------------------------

    cliente_tipo_rows = []

    for cluster in VALID_CLUSTERS:
        if cluster not in df[TARGET_COL].unique():
            continue

        subset = df[df[TARGET_COL] == cluster]

        for col in numeric_features:
            cliente_tipo_rows.append({
                "cluster": cluster,
                "variable": col,
                "mean": subset[col].mean(),
                "median": subset[col].median(),
                "p25": subset[col].quantile(0.25),
                "p75": subset[col].quantile(0.75),
                "missing_rate": subset[col].isna().mean()
            })

    cliente_tipo_df = pd.DataFrame(cliente_tipo_rows)

    # --------------------------------------------------------
    # 3.14 Business insight
    # --------------------------------------------------------

    insights = []

    if not stat_tests_df.empty:
        significant_vars = stat_tests_df.loc[
            stat_tests_df["significant_005"] == True,
            "variable"
        ].tolist()
    else:
        significant_vars = []

    top_important_vars = top_features["feature"].head(20).tolist()

    common_vars = [
        variable
        for variable in top_important_vars
        if variable in numeric_features
        and variable in significant_vars
    ]

    for variable in common_vars[:15]:
        try:
            c30 = cluster_profile.loc[variable, "cluster30"]
            c90 = cluster_profile.loc[variable, "cluster90"]
            c90p = cluster_profile.loc[variable, "cluster90+"]

            if pd.notna(c30) and pd.notna(c90p):
                direction = "higher" if c90p > c30 else "lower"

                insights.append({
                    "variable": variable,
                    "business_insight": (
                        f"{variable}: median value is {direction} in cluster90+ "
                        f"than in cluster30. cluster30={c30:.2f}, "
                        f"cluster90={c90:.2f}, cluster90+={c90p:.2f}."
                    )
                })
        except Exception:
            pass

    business_insights_df = pd.DataFrame(insights)

    # --------------------------------------------------------
    # 3.15 Confronto cluster
    # --------------------------------------------------------

    cluster_comparison = cluster_profile.copy()

    if not cluster_comparison.empty and not stat_tests_df.empty:
        pvals = (
            stat_tests_df
            .drop_duplicates("variable")
            .set_index("variable")["p_value"]
            .to_dict()
        )

        cluster_comparison["p_value"] = (
            cluster_comparison.index.map(pvals)
        )

        cluster_comparison["significant_005"] = (
            cluster_comparison["p_value"] < 0.05
        )

    if (
        not cluster_comparison.empty
        and "abs_delta_90plus_vs_30" in cluster_comparison.columns
    ):
        cluster_comparison = cluster_comparison.sort_values(
            "abs_delta_90plus_vs_30",
            ascending=False
        )

    # --------------------------------------------------------
    # 3.16 Report modelli
    # --------------------------------------------------------

    model_report_rows = []

    for model_name, report in model_reports.items():
        for label, metrics in report.items():
            if isinstance(metrics, dict):
                row = {
                    "model": model_name,
                    "label": label
                }
                row.update(metrics)
                model_report_rows.append(row)
            else:
                model_report_rows.append({
                    "model": model_name,
                    "label": label,
                    "value": metrics
                })

    model_report_df = pd.DataFrame(model_report_rows)

    # --------------------------------------------------------
    # 3.17 Salvataggio modello e metadati
    # --------------------------------------------------------

    joblib.dump(rf_model, MODEL_FILE)

    training_distribution = format_distribution(
        y,
        "training"
    )

    metadata = {
        "model_type": "RandomForestClassifier",
        "target_col": TARGET_COL,
        "valid_clusters": VALID_CLUSTERS,
        "cluster_rank": CLUSTER_RANK,
        "feature_columns": list(X.columns),
        "numeric_features": numeric_features,
        "categorical_features": categorical_features,
        "constant_columns_removed": constant_cols,
        "leakage_columns": LEAKAGE_COLS,
        "id_columns": ID_COLS,
        "training_rows": int(len(df)),
        "random_state": RANDOM_STATE,
        "drift_relative_threshold": DRIFT_RELATIVE_THRESHOLD,
        "drift_std_threshold": DRIFT_STD_THRESHOLD
    }

    with open(METADATA_FILE, "w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)

    with pd.ExcelWriter(
        TRAINING_PROFILE_FILE,
        engine="openpyxl"
    ) as writer:
        training_distribution.to_excel(
            writer,
            sheet_name="training_distribution",
            index=False
        )

        training_numeric_profile.to_excel(
            writer,
            sheet_name="numeric_profile",
            index=False
        )

        cluster_median.to_excel(
            writer,
            sheet_name="cluster_medians"
        )

        cluster_mean.to_excel(
            writer,
            sheet_name="cluster_means"
        )

        cliente_tipo_df.to_excel(
            writer,
            sheet_name="cliente_tipo",
            index=False
        )

    style_excel_file(TRAINING_PROFILE_FILE)

    # --------------------------------------------------------
    # 3.18 Export report completo
    # --------------------------------------------------------

    with pd.ExcelWriter(
        TRAIN_OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        training_distribution.to_excel(
            writer,
            sheet_name="00_cluster_distribution",
            index=False
        )

        model_report_df.to_excel(
            writer,
            sheet_name="01_model_performance",
            index=False
        )

        cv_summary.to_excel(
            writer,
            sheet_name="02_cross_validation",
            index=False
        )

        feature_importance.to_excel(
            writer,
            sheet_name="03_feature_importance",
            index=False
        )

        cluster_comparison.to_excel(
            writer,
            sheet_name="04_cluster_comparison"
        )

        stat_tests_df.to_excel(
            writer,
            sheet_name="05_stat_tests",
            index=False
        )

        cliente_tipo_df.to_excel(
            writer,
            sheet_name="06_cliente_tipo",
            index=False
        )

        score_df.to_excel(
            writer,
            sheet_name="07_scores_clienti",
            index=False
        )

        business_insights_df.to_excel(
            writer,
            sheet_name="08_business_insights",
            index=False
        )

        pd.DataFrame({
            "decision_tree_rules": tree_rules.split("\n")
        }).to_excel(
            writer,
            sheet_name="09_decision_tree_rules",
            index=False
        )

        training_numeric_profile.to_excel(
            writer,
            sheet_name="10_training_profile",
            index=False
        )

        pd.DataFrame({
            "artifact": [
                "Random Forest pipeline",
                "Model metadata",
                "Training profiles"
            ],
            "path": [
                str(MODEL_FILE),
                str(METADATA_FILE),
                str(TRAINING_PROFILE_FILE)
            ]
        }).to_excel(
            writer,
            sheet_name="11_saved_artifacts",
            index=False
        )

        for name, cm in confusion_matrices.items():
            safe_name = name.replace(" ", "_")[:20]

            cm.to_excel(
                writer,
                sheet_name=f"CM_{safe_name}"
            )

    style_excel_file(TRAIN_OUTPUT_FILE)

    print(f"\nFile report esportato: {TRAIN_OUTPUT_FILE}")
    print(f"Pipeline salvata: {MODEL_FILE}")
    print(f"Metadati salvati: {METADATA_FILE}")
    print(f"Profili training salvati: {TRAINING_PROFILE_FILE}")


# ============================================================
# 4. PREDIZIONE SU NUOVO DATASET
# ============================================================

def predict_new_dataset():
    required_files = [
        MODEL_FILE,
        METADATA_FILE,
        TRAINING_PROFILE_FILE
    ]

    missing_artifacts = [
        str(path)
        for path in required_files
        if not path.exists()
    ]

    if missing_artifacts:
        raise FileNotFoundError(
            "Mancano gli artefatti di training:\n- "
            + "\n- ".join(missing_artifacts)
            + "\nEseguire prima MODE = 'TRAIN'."
        )

    model = joblib.load(MODEL_FILE)

    with open(METADATA_FILE, "r", encoding="utf-8") as file:
        metadata = json.load(file)

    feature_columns = metadata["feature_columns"]
    numeric_features = metadata["numeric_features"]

    new_df = pd.read_excel(NEW_DATA_FILE)

    print("Righe nuovo dataset:", len(new_df))
    print("Colonne nuovo dataset:", len(new_df.columns))

    # --------------------------------------------------------
    # 4.1 Target opzionale
    # --------------------------------------------------------

    actual_target = None
    evaluation_mask = None

    if TARGET_COL in new_df.columns:
        normalized_target = normalize_target(new_df[TARGET_COL])

        evaluation_mask = normalized_target.isin(VALID_CLUSTERS)

        if evaluation_mask.any():
            actual_target = normalized_target
            print(
                "\nTarget disponibile per "
                f"{int(evaluation_mask.sum())} osservazioni su {len(new_df)}."
            )
        else:
            print(
                "\nLa colonna target è presente ma non contiene "
                "etichette riconosciute."
            )

    # --------------------------------------------------------
    # 4.2 Allineamento colonne
    # --------------------------------------------------------

    X_new_raw = new_df.drop(
        columns=(
            [TARGET_COL]
            + TARGET_DERIVED_COLS
            + LEAKAGE_COLS
            + ID_COLS
        ),
        errors="ignore"
    )

    missing_features = [
        col for col in feature_columns
        if col not in X_new_raw.columns
    ]

    extra_features = [
        col for col in X_new_raw.columns
        if col not in feature_columns
    ]

    # Le feature mancanti vengono aggiunte come NaN.
    # Il SimpleImputer della pipeline userà le statistiche apprese nel training.
    for col in missing_features:
        X_new_raw[col] = np.nan

    X_new = X_new_raw.reindex(columns=feature_columns)

    column_check_df = pd.DataFrame({
        "feature": sorted(set(feature_columns + list(X_new_raw.columns))),
        "status": [
            (
                "missing_in_new_dataset"
                if feature in missing_features
                else "extra_not_used"
                if feature in extra_features
                else "used"
            )
            for feature in sorted(
                set(feature_columns + list(X_new_raw.columns))
            )
        ]
    })

    print("\nFeature mancanti aggiunte come NaN:", len(missing_features))
    print("Feature extra ignorate:", len(extra_features))

    # --------------------------------------------------------
    # 4.3 Predizione e scoring
    # --------------------------------------------------------

    score_df = build_score_dataframe(
        model=model,
        X=X_new,
        source_df=new_df,
        actual_target=actual_target
    )

    predictions = score_df["predicted_cluster"]

    # Dataset completo con cluster e score.
    output_scored = new_df.copy()
    output_scored["predicted_cluster"] = predictions.values
    output_scored["management_persistence_score"] = (
        score_df["management_persistence_score"].values
    )
    output_scored["score_band"] = score_df["score_band"].values

    for col in [c for c in score_df.columns if c.startswith("prob_")]:
        output_scored[col] = score_df[col].values

    # --------------------------------------------------------
    # 4.4 Confronto distribuzione cluster
    # --------------------------------------------------------

    training_distribution = pd.read_excel(
        TRAINING_PROFILE_FILE,
        sheet_name="training_distribution"
    )

    new_distribution = format_distribution(
        predictions,
        "new"
    )

    distribution_comparison = training_distribution.merge(
        new_distribution,
        on="cluster",
        how="outer"
    ).fillna(0)

    distribution_comparison["percentage_point_change"] = (
        distribution_comparison["new_percentage"]
        - distribution_comparison["training_percentage"]
    )

    distribution_comparison["relative_change"] = (
        distribution_comparison.apply(
            lambda row: safe_relative_change(
                row["new_percentage"],
                row["training_percentage"]
            ),
            axis=1
        )
    )

    # --------------------------------------------------------
    # 4.5 Drift numerico
    # --------------------------------------------------------

    training_numeric_profile = pd.read_excel(
        TRAINING_PROFILE_FILE,
        sheet_name="numeric_profile"
    )

    drift_rows = []

    for col in numeric_features:
        if col not in X_new.columns:
            continue

        series = pd.to_numeric(
            X_new[col],
            errors="coerce"
        )

        training_row = training_numeric_profile.loc[
            training_numeric_profile["variable"] == col
        ]

        if training_row.empty:
            continue

        training_row = training_row.iloc[0]

        new_median = series.median()
        new_mean = series.mean()
        new_std = series.std()
        new_missing_rate = series.isna().mean()

        training_median = training_row["training_median"]
        training_std = training_row["training_std"]
        training_missing_rate = training_row["training_missing_rate"]

        relative_change = safe_relative_change(
            new_median,
            training_median
        )

        if (
            pd.notna(training_std)
            and abs(training_std) > 1e-9
            and pd.notna(new_median)
            and pd.notna(training_median)
        ):
            standardized_shift = (
                new_median - training_median
            ) / training_std
        else:
            standardized_shift = np.nan

        missing_rate_change = (
            new_missing_rate - training_missing_rate
        )

        relative_flag = (
            pd.notna(relative_change)
            and abs(relative_change) >= DRIFT_RELATIVE_THRESHOLD
        )

        standardized_flag = (
            pd.notna(standardized_shift)
            and abs(standardized_shift) >= DRIFT_STD_THRESHOLD
        )

        drift_rows.append({
            "variable": col,
            "training_median": training_median,
            "new_median": new_median,
            "relative_median_change": relative_change,
            "abs_relative_median_change": (
                abs(relative_change)
                if pd.notna(relative_change)
                else np.nan
            ),
            "training_mean": training_row["training_mean"],
            "new_mean": new_mean,
            "training_std": training_std,
            "new_std": new_std,
            "standardized_median_shift": standardized_shift,
            "abs_standardized_median_shift": (
                abs(standardized_shift)
                if pd.notna(standardized_shift)
                else np.nan
            ),
            "training_missing_rate": training_missing_rate,
            "new_missing_rate": new_missing_rate,
            "missing_rate_change": missing_rate_change,
            "relative_threshold": DRIFT_RELATIVE_THRESHOLD,
            "std_threshold": DRIFT_STD_THRESHOLD,
            "drift_flag": relative_flag or standardized_flag
        })

    drift_df = pd.DataFrame(drift_rows)

    if not drift_df.empty:
        drift_df = drift_df.sort_values(
            [
                "drift_flag",
                "abs_standardized_median_shift",
                "abs_relative_median_change"
            ],
            ascending=[False, False, False]
        )

    # --------------------------------------------------------
    # 4.6 Valutazione, se target disponibile
    # --------------------------------------------------------

    evaluation_report_df = pd.DataFrame()
    evaluation_cm_df = pd.DataFrame()

    if actual_target is not None and evaluation_mask is not None:
        valid_idx = evaluation_mask[evaluation_mask].index

        y_true = actual_target.loc[valid_idx]
        y_pred = pd.Series(
            model.predict(X_new.loc[valid_idx]),
            index=valid_idx
        )

        report = classification_report(
            y_true,
            y_pred,
            labels=VALID_CLUSTERS,
            output_dict=True,
            zero_division=0
        )

        report_rows = []

        for label, metrics in report.items():
            if isinstance(metrics, dict):
                row = {"label": label}
                row.update(metrics)
                report_rows.append(row)
            else:
                report_rows.append({
                    "label": label,
                    "value": metrics
                })

        evaluation_report_df = pd.DataFrame(report_rows)

        evaluation_cm_df = pd.DataFrame(
            confusion_matrix(
                y_true,
                y_pred,
                labels=VALID_CLUSTERS
            ),
            index=[f"actual_{c}" for c in VALID_CLUSTERS],
            columns=[f"pred_{c}" for c in VALID_CLUSTERS]
        )

    # --------------------------------------------------------
    # 4.7 Profilo del nuovo dataset per cluster predetto
    # --------------------------------------------------------

    new_profile_source = X_new.copy()
    new_profile_source["predicted_cluster"] = predictions.values

    if numeric_features:
        new_cluster_profile = (
            new_profile_source
            .groupby("predicted_cluster")[numeric_features]
            .median()
            .T
        )

        for cluster in VALID_CLUSTERS:
            if cluster not in new_cluster_profile.columns:
                new_cluster_profile[cluster] = np.nan

        new_cluster_profile = new_cluster_profile[VALID_CLUSTERS]
    else:
        new_cluster_profile = pd.DataFrame()

    portfolio_summary_df = generate_portfolio_summary(
        distribution_comparison=distribution_comparison,
        drift_df=drift_df
    )

    # --------------------------------------------------------
    # 4.8 Export
    # --------------------------------------------------------

    with pd.ExcelWriter(
        PREDICT_OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        output_scored.to_excel(
            writer,
            sheet_name="00_scored_dataset",
            index=False
        )

        score_df.to_excel(
            writer,
            sheet_name="01_scores_probabilities",
            index=False
        )

        distribution_comparison.to_excel(
            writer,
            sheet_name="02_cluster_comparison",
            index=False
        )

        drift_df.to_excel(
            writer,
            sheet_name="03_numeric_drift",
            index=False
        )

        new_cluster_profile.to_excel(
            writer,
            sheet_name="04_new_cluster_profile"
        )

        portfolio_summary_df.to_excel(
            writer,
            sheet_name="05_portfolio_summary",
            index=False
        )

        column_check_df.to_excel(
            writer,
            sheet_name="06_column_check",
            index=False
        )

        if not evaluation_report_df.empty:
            evaluation_report_df.to_excel(
                writer,
                sheet_name="07_model_evaluation",
                index=False
            )

            evaluation_cm_df.to_excel(
                writer,
                sheet_name="08_confusion_matrix"
            )

    style_excel_file(PREDICT_OUTPUT_FILE)

    print(f"\nFile nuovo dataset esportato: {PREDICT_OUTPUT_FILE}")

    if missing_features:
        print(
            "\nATTENZIONE: alcune feature del training mancavano nel nuovo file. "
            "Sono state valorizzate come NaN e gestite dall'imputer:"
        )
        print(missing_features)

    if extra_features:
        print(
            "\nColonne presenti nel nuovo dataset ma non utilizzate dal modello:"
        )
        print(extra_features)


# ============================================================
# 5. MAIN
# ============================================================

def main():
    selected_mode = MODE.strip().upper()

    if selected_mode == "TRAIN":
        train_model()

    elif selected_mode == "PREDICT":
        predict_new_dataset()

    else:
        raise ValueError(
            "MODE deve essere impostato su 'TRAIN' oppure 'PREDICT'."
        )


if __name__ == "__main__":
    main()
